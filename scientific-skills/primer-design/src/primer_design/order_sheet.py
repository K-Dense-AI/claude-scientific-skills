#!/usr/bin/env python3
"""
Primer Order Sheet Generator
==============================
프라이머 설계 결과를 Macrogen 주문서 형식으로 변환.

지원 출력:
  - XLSX (Macrogen Order + Summary + QC 시트)
  - CSV  (UTF-8 BOM, 한글 Excel 호환)
  - Markdown
  - pandas DataFrame
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from enum import Enum
from pathlib import Path

import pandas as pd


# ── Constants ──────────────────────────────────────────────────────────────

COST_PER_BASE_KRW = 400    # 25 nmol scale
MIN_PRIMER_COST_KRW = 5000

# Macrogen 주문서 매핑
_SCALE_TO_UMOL = {
    "25 nmol": 0.025,
    "50 nmol": 0.05,
    "100 nmol": 0.1,
    "1 umol": 1.0,
}

_PURIFICATION_TO_MACROGEN = {
    "Desalting": "MOPC",
    "PAGE": "PAGE",
    "HPLC": "HPLC",
}


# ── Enums ──────────────────────────────────────────────────────────────────

class PrimerScale(Enum):
    NMOL_25 = "25 nmol"
    NMOL_50 = "50 nmol"
    NMOL_100 = "100 nmol"
    UMOL_1 = "1 umol"


class Purification(Enum):
    DESALTING = "Desalting"
    PAGE = "PAGE"
    HPLC = "HPLC"


# ── Data class ─────────────────────────────────────────────────────────────

@dataclass
class PrimerEntry:
    name: str
    sequence: str       # 5'->3', uppercase
    length: int
    scale: PrimerScale
    purification: Purification
    tm: float | None
    gc: float | None
    qc_verdict: str | None
    experiment: str
    direction: str      # "F" or "R"
    notes: str


# ── Order Sheet ────────────────────────────────────────────────────────────

class PrimerOrderSheet:
    """프라이머 주문서 생성기.

    iPCR 설계 결과(dict)를 받아 Macrogen 주문서 형식의 출력을 생성한다.
    """

    def __init__(
        self,
        project_name: str = "primer_order",
        default_scale: PrimerScale = PrimerScale.NMOL_25,
        default_purification: Purification = Purification.DESALTING,
    ):
        self.project_name = project_name
        self.default_scale = default_scale
        self.default_purification = default_purification
        self.entries: list[PrimerEntry] = []

    # ── Add from design result ─────────────────────────────────────────

    def add_from_design_result(
        self,
        result: dict,
        experiment: str = "",
        name_prefix: str = "iPCR",
        parent: str = "",
        mutation: str = "",
    ) -> tuple[PrimerEntry, PrimerEntry]:
        """설계 결과 dict에서 F/R 프라이머 엔트리 생성.

        Naming priority:
          1. parent + mutation 제공 시: iPCR_{parent}_{mutation}_F/R
          2. result에 re_5prime, re_3prime 키 존재 시:
             {name_prefix}_{re_5prime}_{re_3prime}_F/R
          3. 기본: {name_prefix}_{seq_number}_F/R
        """
        # Determine names
        if parent and mutation:
            base_name = f"iPCR_{parent}_{mutation}"
        elif "re_5prime" in result and "re_3prime" in result:
            base_name = f"{name_prefix}_{result['re_5prime']}_{result['re_3prime']}"
        else:
            seq_number = (len(self.entries) // 2) + 1
            base_name = f"{name_prefix}_{seq_number:03d}"

        f_name = f"{base_name}_F"
        r_name = f"{base_name}_R"

        # Extract QC verdicts
        f_qc_verdict = None
        if isinstance(result.get("f_qc"), dict):
            f_qc_verdict = result["f_qc"].get("verdict")
        r_qc_verdict = None
        if isinstance(result.get("r_qc"), dict):
            r_qc_verdict = result["r_qc"].get("verdict")

        f_seq = result["f_full"].upper()
        r_seq = result["r_full"].upper()

        f_entry = PrimerEntry(
            name=f_name,
            sequence=f_seq,
            length=len(f_seq),
            scale=self.default_scale,
            purification=self.default_purification,
            tm=result.get("f_tm"),
            gc=result.get("f_gc"),
            qc_verdict=f_qc_verdict,
            experiment=experiment,
            direction="F",
            notes="",
        )
        r_entry = PrimerEntry(
            name=r_name,
            sequence=r_seq,
            length=len(r_seq),
            scale=self.default_scale,
            purification=self.default_purification,
            tm=result.get("r_tm"),
            gc=result.get("r_gc"),
            qc_verdict=r_qc_verdict,
            experiment=experiment,
            direction="R",
            notes="",
        )

        self.entries.append(f_entry)
        self.entries.append(r_entry)
        return f_entry, r_entry

    # ── Add custom primer ──────────────────────────────────────────────

    def add_custom_primer(
        self,
        name: str,
        sequence: str,
        experiment: str = "",
        direction: str = "",
        scale: PrimerScale | None = None,
        purification: Purification | None = None,
        tm: float | None = None,
        gc: float | None = None,
        notes: str = "",
    ) -> PrimerEntry:
        """사용자 정의 프라이머 추가."""
        seq_upper = sequence.upper()
        entry = PrimerEntry(
            name=name,
            sequence=seq_upper,
            length=len(seq_upper),
            scale=scale if scale is not None else self.default_scale,
            purification=purification if purification is not None else self.default_purification,
            tm=tm,
            gc=gc,
            qc_verdict=None,
            experiment=experiment,
            direction=direction,
            notes=notes,
        )
        self.entries.append(entry)
        return entry

    # ── Batch add ──────────────────────────────────────────────────────

    def add_batch_from_results(
        self,
        results: list[dict],
        name_prefix: str = "iPCR",
    ) -> list[tuple[PrimerEntry, PrimerEntry]]:
        """여러 설계 결과를 일괄 추가."""
        pairs = []
        for result in results:
            pair = self.add_from_design_result(result, name_prefix=name_prefix)
            pairs.append(pair)
        return pairs

    # ── Deduplication ──────────────────────────────────────────────────

    def deduplicate(self) -> list[tuple[str, str]]:
        """서열 중복 제거. 첫 번째 엔트리를 유지하고 나머지를 제거.

        Returns
        -------
        list of (removed_name, kept_name)
        """
        seen: dict[str, str] = {}  # sequence -> name of first occurrence
        removed: list[tuple[str, str]] = []
        unique_entries: list[PrimerEntry] = []

        for entry in self.entries:
            seq_key = entry.sequence.upper()
            if seq_key in seen:
                removed.append((entry.name, seen[seq_key]))
            else:
                seen[seq_key] = entry.name
                unique_entries.append(entry)

        self.entries = unique_entries
        return removed

    # ── Summary ────────────────────────────────────────────────────────

    def summary(self) -> dict:
        """주문서 요약 통계."""
        total_primers = len(self.entries)
        total_length_nt = sum(e.length for e in self.entries)

        estimated_cost_krw = 0
        for entry in self.entries:
            cost = entry.length * COST_PER_BASE_KRW
            estimated_cost_krw += max(cost, MIN_PRIMER_COST_KRW)

        unique_experiments = set()
        for entry in self.entries:
            if entry.experiment:
                unique_experiments.add(entry.experiment)

        scale_counts: dict[str, int] = {}
        for entry in self.entries:
            key = entry.scale.value
            scale_counts[key] = scale_counts.get(key, 0) + 1

        purification_counts: dict[str, int] = {}
        for entry in self.entries:
            key = entry.purification.value
            purification_counts[key] = purification_counts.get(key, 0) + 1

        return {
            "total_primers": total_primers,
            "total_length_nt": total_length_nt,
            "estimated_cost_krw": estimated_cost_krw,
            "unique_experiments": sorted(unique_experiments),
            "scale_counts": scale_counts,
            "purification_counts": purification_counts,
        }

    # ── Export: XLSX ───────────────────────────────────────────────────

    def to_xlsx(self, output_path: str | Path | None = None) -> Path:
        """Macrogen 주문서 XLSX 생성.

        Sheet1 "Macrogen Order": No., Primer Name, Sequence, Scale, etc.
        Sheet2 "Summary": 요약 통계
        Sheet3 "QC": Tm, GC%, QC Verdict
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        if output_path is None:
            output_path = self._generate_filename(Path.cwd(), "xlsx")
        else:
            output_path = Path(output_path)

        wb = openpyxl.Workbook()

        # ── Sheet 1: Macrogen Order ────────────────────────────────────
        ws_order = wb.active
        ws_order.title = "Macrogen Order"

        headers = [
            "No.", "Primer Name", "Sequence (5'->3')", "Scale",
            "Purification", "Length (nt)", "Notes",
        ]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws_order.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, entry in enumerate(self.entries, 2):
            ws_order.cell(row=row_idx, column=1, value=row_idx - 1)
            ws_order.cell(row=row_idx, column=2, value=entry.name)
            seq_cell = ws_order.cell(row=row_idx, column=3, value=entry.sequence)
            seq_cell.font = Font(name="Consolas", size=10)
            ws_order.cell(row=row_idx, column=4, value=entry.scale.value)
            ws_order.cell(row=row_idx, column=5, value=entry.purification.value)
            ws_order.cell(row=row_idx, column=6, value=entry.length)
            ws_order.cell(row=row_idx, column=7, value=entry.notes)

        # Column widths
        col_widths = [6, 25, 60, 12, 12, 12, 20]
        for i, width in enumerate(col_widths, 1):
            ws_order.column_dimensions[get_column_letter(i)].width = width

        # ── Sheet 2: Summary ───────────────────────────────────────────
        ws_summary = wb.create_sheet("Summary")
        summary_data = self.summary()

        summary_rows = [
            ("Total Primers", summary_data["total_primers"]),
            ("Total Length (nt)", summary_data["total_length_nt"]),
            ("Estimated Cost (KRW)", f"{summary_data['estimated_cost_krw']:,}"),
            ("Unique Experiments", ", ".join(summary_data["unique_experiments"]) or "-"),
        ]

        ws_summary.cell(row=1, column=1, value="Metric").font = Font(bold=True)
        ws_summary.cell(row=1, column=2, value="Value").font = Font(bold=True)

        for row_idx, (metric, value) in enumerate(summary_rows, 2):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)

        # Scale counts
        row_offset = len(summary_rows) + 3
        ws_summary.cell(row=row_offset, column=1, value="Scale").font = Font(bold=True)
        ws_summary.cell(row=row_offset, column=2, value="Count").font = Font(bold=True)
        for i, (scale, count) in enumerate(summary_data["scale_counts"].items(), 1):
            ws_summary.cell(row=row_offset + i, column=1, value=scale)
            ws_summary.cell(row=row_offset + i, column=2, value=count)

        # Purification counts
        row_offset2 = row_offset + len(summary_data["scale_counts"]) + 2
        ws_summary.cell(row=row_offset2, column=1, value="Purification").font = Font(bold=True)
        ws_summary.cell(row=row_offset2, column=2, value="Count").font = Font(bold=True)
        for i, (pur, count) in enumerate(summary_data["purification_counts"].items(), 1):
            ws_summary.cell(row=row_offset2 + i, column=1, value=pur)
            ws_summary.cell(row=row_offset2 + i, column=2, value=count)

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 30

        # ── Sheet 3: QC ────────────────────────────────────────────────
        ws_qc = wb.create_sheet("QC")

        qc_headers = ["Primer Name", "Tm", "GC%", "QC Verdict", "Experiment"]
        for col_idx, header in enumerate(qc_headers, 1):
            cell = ws_qc.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        verdict_fills = {
            "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "WARNING": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }

        for row_idx, entry in enumerate(self.entries, 2):
            ws_qc.cell(row=row_idx, column=1, value=entry.name)
            ws_qc.cell(row=row_idx, column=2, value=entry.tm)
            ws_qc.cell(row=row_idx, column=3, value=entry.gc)
            verdict_cell = ws_qc.cell(row=row_idx, column=4, value=entry.qc_verdict or "-")
            if entry.qc_verdict in verdict_fills:
                verdict_cell.fill = verdict_fills[entry.qc_verdict]
            ws_qc.cell(row=row_idx, column=5, value=entry.experiment)

        qc_col_widths = [25, 10, 10, 15, 25]
        for i, width in enumerate(qc_col_widths, 1):
            ws_qc.column_dimensions[get_column_letter(i)].width = width

        wb.save(str(output_path))
        return output_path

    # ── Export: Macrogen Oligo Order ──────────────────────────────────

    def to_macrogen_oligo(self, output_path: str | Path | None = None) -> Path:
        """Macrogen Oligo 주문서 형식 (.xlsx) 생성.

        Macrogen 홈페이지 업로드 호환 형식:
          No. | Oligo Name | 5` - Oligo Seq - 3` | Amount | Purification

        Amount: umol 단위 (0.025 = 25 nmol)
        Purification: MOPC (desalting), PAGE, HPLC
        빈 행 포함 총 1000행 (Macrogen 템플릿 호환).
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font

        if output_path is None:
            output_path = self._generate_filename(Path.cwd(), "xlsx")
        else:
            output_path = Path(output_path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet"

        # 헤더 (Macrogen 템플릿과 동일)
        headers = ["No.", "Oligo Name", "5` - Oligo Seq - 3`", "Amount", "Purification"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 데이터 + 빈 행 (총 1000행)
        for row_num in range(1, 1001):
            row_idx = row_num + 1  # 헤더가 1행
            ws.cell(row=row_idx, column=1, value=row_num)

            if row_num <= len(self.entries):
                entry = self.entries[row_num - 1]
                ws.cell(row=row_idx, column=2, value=entry.name)
                seq_cell = ws.cell(row=row_idx, column=3, value=entry.sequence)
                seq_cell.font = Font(name="Consolas", size=10)
                amount = _SCALE_TO_UMOL.get(entry.scale.value, 0.025)
                ws.cell(row=row_idx, column=4, value=amount)
                pur = _PURIFICATION_TO_MACROGEN.get(entry.purification.value, "MOPC")
                ws.cell(row=row_idx, column=5, value=pur)
            else:
                ws.cell(row=row_idx, column=2, value="")
                ws.cell(row=row_idx, column=3, value="")
                ws.cell(row=row_idx, column=4, value="")
                ws.cell(row=row_idx, column=5, value="")

        # 컬럼 너비
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 14

        wb.save(str(output_path))
        return output_path

    # ── Export: Macrogen Sequencing Order ─────────────────────────────

    def to_macrogen_seq(
        self,
        sample_primer_pairs: list[dict],
        output_path: str | Path | None = None,
    ) -> Path:
        """Macrogen Standard Sequencing 주문서 형식 (.xlsx) 생성.

        Macrogen 홈페이지 업로드 호환 형식:
          # | Sample Name | Primer Name | Sample Concentration (ng/ul) |
          Plate Name | Well Position | Product Size(bp) | Target Size(bp) |
          Primer Sequence(5 to 3) | Primer Concentration (pmol/ul)

        Parameters
        ----------
        sample_primer_pairs : list[dict]
            각 항목:
            {
                "sample_name": str,                      # 필수
                "primer_name": str,                      # 필수
                "sample_conc": float | None,             # ng/ul
                "plate_name": str,                       # optional
                "well_position": str,                    # optional
                "product_size": int | None,              # bp
                "target_size": int | None,               # bp
                "primer_seq": str,                       # optional (5'→3')
                "primer_conc": float | None,             # pmol/ul
            }
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font

        if output_path is None:
            output_path = self._generate_filename(Path.cwd(), "xlsx")
        else:
            output_path = Path(output_path)

        wb = openpyxl.Workbook()

        # ── Sheet1: Order ─────────────────────────────────────────────
        ws = wb.active
        ws.title = "Sheet1"

        # Row 1: 주의사항
        ws.cell(
            row=1, column=1,
            value="     ※ Only English Alphabet (either capital small letters), "
                  "digit 0~9, a hypen (-) or under bar (_) is allowed "
                  "without any blanks.",
        )

        # Row 2: 그룹 헤더
        ws.cell(row=2, column=1, value="#")
        ws.cell(row=2, column=2, value="Reaction Information")
        ws.cell(row=2, column=4, value="Sample Information")
        ws.cell(row=2, column=9, value="Primer Information")
        for col in [1, 2, 4, 9]:
            ws.cell(row=2, column=col).font = Font(bold=True)

        # Row 3: 컬럼 헤더
        col_headers = [
            "#", "Sample Name *", "Primer Name *",
            "Sample Concentration (ng/ul)", "Plate Name", "Well Position",
            "Product Size(bp)", "Target Size(bp)",
            "Primer Sequence(5 to 3)", "Primer Concentration (pmol/ul)",
        ]
        for col_idx, header in enumerate(col_headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # 데이터 행 (row 4부터)
        total_rows = max(len(sample_primer_pairs), 1000)
        for row_num in range(1, total_rows + 1):
            row_idx = row_num + 3
            ws.cell(row=row_idx, column=1, value=row_num)

            if row_num <= len(sample_primer_pairs):
                sp = sample_primer_pairs[row_num - 1]
                ws.cell(row=row_idx, column=2, value=sp.get("sample_name", ""))
                ws.cell(row=row_idx, column=3, value=sp.get("primer_name", ""))
                ws.cell(row=row_idx, column=4, value=sp.get("sample_conc"))
                ws.cell(row=row_idx, column=5, value=sp.get("plate_name", ""))
                ws.cell(row=row_idx, column=6, value=sp.get("well_position", ""))
                ws.cell(row=row_idx, column=7, value=sp.get("product_size"))
                ws.cell(row=row_idx, column=8, value=sp.get("target_size"))
                seq_cell = ws.cell(row=row_idx, column=9, value=sp.get("primer_seq", ""))
                seq_cell.font = Font(name="Consolas", size=10)
                ws.cell(row=row_idx, column=10, value=sp.get("primer_conc"))

        # 컬럼 너비
        widths = [5, 25, 25, 15, 12, 12, 12, 12, 45, 15]
        for i, w in enumerate(widths, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Sheet2: Reference ─────────────────────────────────────────
        ws2 = wb.create_sheet("Sheet2")
        ws2.cell(row=2, column=1, value="Product Size(bp)").font = Font(bold=True)
        ws2.cell(row=2, column=2, value="Sample Align").font = Font(bold=True)
        ws2.cell(row=3, column=1, value="600bp Over")
        ws2.cell(row=3, column=2, value="Vertical")
        ws2.cell(row=4, column=1, value="600bp Less")
        ws2.cell(row=4, column=2, value="Horizontal")

        wb.save(str(output_path))
        return output_path

    # ── Export: CSV ────────────────────────────────────────────────────

    def to_csv(self, output_path: str | Path | None = None) -> Path:
        """UTF-8 BOM CSV (한글 Excel 호환)."""
        if output_path is None:
            output_path = self._generate_filename(Path.cwd(), "csv")
        else:
            output_path = Path(output_path)

        df = self.to_dataframe()
        df.to_csv(str(output_path), index=False, encoding="utf-8-sig")
        return output_path

    # ── Export: Markdown ───────────────────────────────────────────────

    def to_markdown(self, output_path: str | Path | None = None) -> Path:
        """Markdown 주문서 생성."""
        if output_path is None:
            output_path = self._generate_filename(Path.cwd(), "md")
        else:
            output_path = Path(output_path)

        lines: list[str] = []
        lines.append(f"# Primer Order: {self.project_name}")
        lines.append(f"")
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        lines.append(f"")

        # Order table
        lines.append("## Order List")
        lines.append("")
        lines.append("| No. | Primer Name | Sequence (5'->3') | Scale | Purification | Length (nt) | Notes |")
        lines.append("|-----|-------------|-------------------|-------|--------------|------------|-------|")
        for i, entry in enumerate(self.entries, 1):
            lines.append(
                f"| {i} | {entry.name} | `{entry.sequence}` | "
                f"{entry.scale.value} | {entry.purification.value} | "
                f"{entry.length} | {entry.notes} |"
            )
        lines.append("")

        # QC table
        lines.append("## QC Summary")
        lines.append("")
        lines.append("| Primer Name | Tm | GC% | QC Verdict | Experiment |")
        lines.append("|-------------|----|-----|------------|------------|")
        for entry in self.entries:
            tm_str = f"{entry.tm:.1f}" if entry.tm is not None else "-"
            gc_str = f"{entry.gc:.1f}" if entry.gc is not None else "-"
            lines.append(
                f"| {entry.name} | {tm_str} | {gc_str} | "
                f"{entry.qc_verdict or '-'} | {entry.experiment} |"
            )
        lines.append("")

        # Summary
        summary_data = self.summary()
        lines.append("## Summary")
        lines.append("")
        lines.append(f"- Total primers: {summary_data['total_primers']}")
        lines.append(f"- Total length: {summary_data['total_length_nt']} nt")
        lines.append(f"- Estimated cost: {summary_data['estimated_cost_krw']:,} KRW")
        if summary_data["unique_experiments"]:
            lines.append(f"- Experiments: {', '.join(summary_data['unique_experiments'])}")
        lines.append("")

        output_path.write_text("\n".join(lines), encoding="utf-8")
        return output_path

    # ── Export: DataFrame ──────────────────────────────────────────────

    def to_dataframe(self) -> pd.DataFrame:
        """pandas DataFrame 변환."""
        rows = []
        for i, entry in enumerate(self.entries, 1):
            rows.append({
                "No.": i,
                "Primer Name": entry.name,
                "Sequence (5'->3')": entry.sequence,
                "Scale": entry.scale.value,
                "Purification": entry.purification.value,
                "Length (nt)": entry.length,
                "Tm": entry.tm,
                "GC%": entry.gc,
                "QC Verdict": entry.qc_verdict or "-",
                "Experiment": entry.experiment,
                "Direction": entry.direction,
                "Notes": entry.notes,
            })
        return pd.DataFrame(rows)

    # ── Filename generation ────────────────────────────────────────────

    def _generate_filename(self, output_dir: Path, ext: str) -> Path:
        date_str = datetime.now().strftime("%Y%m%d")
        existing = sorted(output_dir.glob(f"{self.project_name}_{date_str}_*_order.{ext}"))
        seq = len(existing) + 1
        return output_dir / f"{self.project_name}_{date_str}_{seq:03d}_order.{ext}"


# ── Tests ──────────────────────────────────────────────────────────────────

def _run_tests():
    """PrimerOrderSheet 기능 테스트."""
    import tempfile

    sep = "=" * 70

    mock_result_1 = {
        "f_full": "ATGCGTAACCTGGCGATCAAGCTG",
        "r_full": "CAGCTTGATCGCCAGGTTACGCAT",
        "f_tm": 62.5, "r_tm": 62.5,
        "f_gc": 54.2, "r_gc": 54.2,
        "f_qc": {"verdict": "PASS"}, "r_qc": {"verdict": "PASS"},
    }
    mock_result_2 = {
        "f_full": "GGATCCATGAAAGCTGCCATTGTTCTG",
        "r_full": "CTCGAGTTATTCAACATCGGTCGC",
        "f_tm": 63.1, "r_tm": 61.8,
        "f_gc": 51.9, "r_gc": 50.0,
        "f_qc": {"verdict": "PASS"}, "r_qc": {"verdict": "WARNING"},
        "re_5prime": "BamHI", "re_3prime": "XhoI",
    }

    tmpdir = Path(tempfile.mkdtemp())

    # ── Test 1: Add from design results + XLSX export ──────────────────
    print(f"\n{sep}\n  Test 1: Add from design results + XLSX export\n{sep}")

    sheet = PrimerOrderSheet(project_name="test_order")

    f1, r1 = sheet.add_from_design_result(
        mock_result_1,
        experiment="site-directed mutagenesis",
        parent="UDH_WT",
        mutation="D280N",
    )
    assert f1.name == "iPCR_UDH_WT_D280N_F", f"Expected iPCR_UDH_WT_D280N_F, got {f1.name}"
    assert r1.name == "iPCR_UDH_WT_D280N_R", f"Expected iPCR_UDH_WT_D280N_R, got {r1.name}"
    assert f1.sequence == "ATGCGTAACCTGGCGATCAAGCTG"
    assert f1.tm == 62.5
    assert f1.qc_verdict == "PASS"
    print(f"  F: {f1.name} -> {f1.sequence} ({f1.length} nt)")
    print(f"  R: {r1.name} -> {r1.sequence} ({r1.length} nt)")

    f2, r2 = sheet.add_from_design_result(
        mock_result_2,
        experiment="restriction cloning",
    )
    assert f2.name == "iPCR_BamHI_XhoI_F", f"Expected iPCR_BamHI_XhoI_F, got {f2.name}"
    assert r2.name == "iPCR_BamHI_XhoI_R", f"Expected iPCR_BamHI_XhoI_R, got {r2.name}"
    assert r2.qc_verdict == "WARNING"
    print(f"  F: {f2.name} -> {f2.sequence} ({f2.length} nt)")
    print(f"  R: {r2.name} -> {r2.sequence} ({r2.length} nt)")

    xlsx_path = sheet.to_xlsx(tmpdir / "test_order.xlsx")
    assert xlsx_path.exists(), f"XLSX not created: {xlsx_path}"
    print(f"  XLSX: {xlsx_path}")
    print("  -> Test 1 PASSED")

    # ── Test 2: Add custom primer ──────────────────────────────────────
    print(f"\n{sep}\n  Test 2: Add custom primer\n{sep}")

    custom = sheet.add_custom_primer(
        name="T7_promoter_F",
        sequence="TAATACGACTCACTATAGGG",
        experiment="sequencing",
        direction="F",
        notes="standard sequencing primer",
    )
    assert custom.name == "T7_promoter_F"
    assert custom.length == 20
    assert custom.sequence == "TAATACGACTCACTATAGGG"
    assert custom.notes == "standard sequencing primer"
    print(f"  Custom: {custom.name} -> {custom.sequence} ({custom.length} nt)")
    print("  -> Test 2 PASSED")

    # ── Test 3: Deduplication ──────────────────────────────────────────
    print(f"\n{sep}\n  Test 3: Deduplication\n{sep}")

    sheet_dup = PrimerOrderSheet(project_name="dup_test")
    sheet_dup.add_custom_primer(name="primer_A", sequence="ATGCGTAACCTGGCG")
    sheet_dup.add_custom_primer(name="primer_B", sequence="atgcgtaacctggcg")  # same, lowercase
    sheet_dup.add_custom_primer(name="primer_C", sequence="GGATCCATGAAAGCT")
    assert len(sheet_dup.entries) == 3

    removed = sheet_dup.deduplicate()
    assert len(sheet_dup.entries) == 2, f"Expected 2 entries after dedup, got {len(sheet_dup.entries)}"
    assert len(removed) == 1
    assert removed[0] == ("primer_B", "primer_A"), f"Expected ('primer_B', 'primer_A'), got {removed[0]}"
    print(f"  Before: 3 primers, After: {len(sheet_dup.entries)} primers")
    print(f"  Removed: {removed}")
    print("  -> Test 3 PASSED")

    # ── Test 4: Summary statistics ─────────────────────────────────────
    print(f"\n{sep}\n  Test 4: Summary statistics\n{sep}")

    summary = sheet.summary()
    assert summary["total_primers"] == 5, f"Expected 5, got {summary['total_primers']}"
    expected_length = 24 + 24 + 27 + 24 + 20
    assert summary["total_length_nt"] == expected_length, (
        f"Expected {expected_length}, got {summary['total_length_nt']}"
    )

    # Cost: each primer = max(length * 400, 5000)
    expected_cost = (
        max(24 * 400, 5000)   # iPCR_UDH_WT_D280N_F
        + max(24 * 400, 5000) # iPCR_UDH_WT_D280N_R
        + max(27 * 400, 5000) # iPCR_BamHI_XhoI_F
        + max(24 * 400, 5000) # iPCR_BamHI_XhoI_R
        + max(20 * 400, 5000) # T7_promoter_F
    )
    assert summary["estimated_cost_krw"] == expected_cost, (
        f"Expected {expected_cost}, got {summary['estimated_cost_krw']}"
    )
    assert summary["scale_counts"]["25 nmol"] == 5
    assert summary["purification_counts"]["Desalting"] == 5
    assert set(summary["unique_experiments"]) == {"site-directed mutagenesis", "restriction cloning", "sequencing"}

    print(f"  Total primers: {summary['total_primers']}")
    print(f"  Total length: {summary['total_length_nt']} nt")
    print(f"  Estimated cost: {summary['estimated_cost_krw']:,} KRW")
    print(f"  Experiments: {summary['unique_experiments']}")
    print(f"  Scales: {summary['scale_counts']}")
    print(f"  Purification: {summary['purification_counts']}")
    print("  -> Test 4 PASSED")

    # ── Test 5: Markdown export ────────────────────────────────────────
    print(f"\n{sep}\n  Test 5: Markdown export\n{sep}")

    md_path = sheet.to_markdown(tmpdir / "test_order.md")
    assert md_path.exists(), f"Markdown not created: {md_path}"
    md_content = md_path.read_text(encoding="utf-8")
    assert "# Primer Order: test_order" in md_content
    assert "iPCR_UDH_WT_D280N_F" in md_content
    assert "iPCR_BamHI_XhoI_F" in md_content
    assert "T7_promoter_F" in md_content
    assert "## QC Summary" in md_content
    assert "## Summary" in md_content
    print(f"  Markdown: {md_path}")
    print(f"  Content length: {len(md_content)} chars")
    print("  -> Test 5 PASSED")

    # ── Test 6: CSV export (UTF-8 BOM) ─────────────────────────────────
    print(f"\n{sep}\n  Test 6: CSV export (UTF-8 BOM check)\n{sep}")

    csv_path = sheet.to_csv(tmpdir / "test_order.csv")
    assert csv_path.exists(), f"CSV not created: {csv_path}"
    csv_bytes = csv_path.read_bytes()
    assert csv_bytes[:3] == b"\xef\xbb\xbf", "CSV missing UTF-8 BOM"
    csv_content = csv_bytes.decode("utf-8-sig")
    assert "iPCR_UDH_WT_D280N_F" in csv_content
    assert "T7_promoter_F" in csv_content
    print(f"  CSV: {csv_path}")
    print(f"  UTF-8 BOM: present")
    print(f"  Lines: {len(csv_content.splitlines())}")
    print("  -> Test 6 PASSED")

    # ── Test 7: Batch add + DataFrame ──────────────────────────────────
    print(f"\n{sep}\n  Test 7: Batch add + DataFrame\n{sep}")

    sheet_batch = PrimerOrderSheet(project_name="batch_test")
    pairs = sheet_batch.add_batch_from_results([mock_result_1, mock_result_2])
    assert len(pairs) == 2
    assert len(sheet_batch.entries) == 4

    df = sheet_batch.to_dataframe()
    assert len(df) == 4
    assert list(df.columns) == [
        "No.", "Primer Name", "Sequence (5'->3')", "Scale", "Purification",
        "Length (nt)", "Tm", "GC%", "QC Verdict", "Experiment", "Direction", "Notes",
    ]
    print(f"  Batch pairs: {len(pairs)}")
    print(f"  DataFrame shape: {df.shape}")
    print(f"  Columns: {list(df.columns)}")
    print("  -> Test 7 PASSED")

    # ── Test 8: Auto filename generation ───────────────────────────────
    print(f"\n{sep}\n  Test 8: Auto filename generation\n{sep}")

    sheet_auto = PrimerOrderSheet(project_name="autoname")
    path1 = sheet_auto._generate_filename(tmpdir, "xlsx")
    assert "autoname_" in path1.name
    assert "_001_order.xlsx" in path1.name
    print(f"  Auto filename: {path1.name}")
    print("  -> Test 8 PASSED")

    print(f"\n{sep}\n  All tests PASSED\n{sep}")
    print(f"  Temp dir: {tmpdir}")


if __name__ == "__main__":
    _run_tests()
