"""FBE 시약목록 수집 - 주간 브리핑용
시트: https://docs.google.com/spreadsheets/d/1_Vn0HKqmCpqU6YAH_Odzcv74H7Y_t4sG6CPVS7wDwho
"""
import json
import io
import sys
import openpyxl
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from google_auth import download_sheet_xlsx

TEMP_DIR = Path(r"C:\Users\Jahyun\lab-analyses\temp")
OUTPUT_FILE = TEMP_DIR / "briefing_reagents.json"
FILE_ID = "1_Vn0HKqmCpqU6YAH_Odzcv74H7Y_t4sG6CPVS7wDwho"

CURRENT_YEAR = datetime.now().year
INSPECTION_KEYWORDS = ["검수", "확인", str(CURRENT_YEAR), "inspection"]


def find_inspection_col(header_row):
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        val = str(cell).strip()
        for kw in INSPECTION_KEYWORDS:
            if kw in val:
                return i
    return None


def collect():
    results = {"sheets": {}, "collected_at": str(datetime.now().date()), "error": None}

    try:
        xlsx_bytes = download_sheet_xlsx(FILE_ID)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        results["error"] = str(e)
        TEMP_DIR.mkdir(parents=True, exist_ok=True)
        OUTPUT_FILE.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[Reagents] ERROR: {e}")
        sys.exit(1)

    total_unchecked = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header = rows[0]
        inspection_col = find_inspection_col(header)
        name_col = 0

        unchecked = []
        for row in rows[1:]:
            if not row or not row[name_col]:
                continue
            name = str(row[name_col]).strip()
            if not name or name == "None":
                continue

            if inspection_col is not None:
                val = row[inspection_col]
                if val is None or str(val).strip() in ("", "None"):
                    unchecked.append({
                        "name": name,
                        "row_data": [str(c) if c is not None else "" for c in row[:6]]
                    })

        results["sheets"][sheet_name] = {
            "unchecked_count": len(unchecked),
            "unchecked": unchecked[:10],
            "inspection_col_found": inspection_col is not None
        }
        total_unchecked += len(unchecked)

    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[Reagents] done: {len(wb.sheetnames)} sheets, unchecked={total_unchecked}")
    print(f"saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    collect()
