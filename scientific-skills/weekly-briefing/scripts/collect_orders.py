"""FBE 주문목록 수집 - 주간 브리핑용
시트: https://docs.google.com/spreadsheets/d/1MD9pugKxYEjk8NX0Dr-YSHAigSFfSmOuQbr02qNQeGc
"""
import json
import io
import sys
import openpyxl
from datetime import datetime, timedelta
from pathlib import Path

# Windows cp949 환경에서 한글 출력 시 깨짐 방지
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent))
from google_auth import download_sheet_xlsx

TEMP_DIR = Path(r"C:\Users\Jahyun\.claude\briefing_temp")
OUTPUT_FILE = TEMP_DIR / "briefing_orders.json"
FILE_ID = "1MD9pugKxYEjk8NX0Dr-YSHAigSFfSmOuQbr02qNQeGc"

CURRENT_YEAR = datetime.now().year
TARGET_SHEET_NAMES = [f"시약 및 소모품_{CURRENT_YEAR}", f"시약및소모품_{CURRENT_YEAR}", str(CURRENT_YEAR)]

# 열 인덱스 (1-based)
COL_ITEM = 1        # A: 품명
COL_REQUESTER = 3   # C: 요청자
COL_DATE_REQ = 6    # F: 주문요청일
COL_STATUS = 11     # K: 주문 상태

def collect():
    results = {"pending": [], "recent_2weeks": [], "collected_at": str(datetime.now().date()), "error": None}

    try:
        xlsx_bytes = download_sheet_xlsx(FILE_ID)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        results["error"] = str(e)
        TEMP_DIR.mkdir(parents=True, exist_ok=True)
        OUTPUT_FILE.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[Orders] ERROR: {e}")
        sys.exit(1)

    # 해당 연도 시트 찾기
    ws = None
    for name in TARGET_SHEET_NAMES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        # 연도가 포함된 시트 이름 검색
        for sname in wb.sheetnames:
            if str(CURRENT_YEAR) in sname:
                ws = wb[sname]
                break
    if ws is None:
        ws = wb.active  # fallback

    cutoff_2w = datetime.now() - timedelta(days=14)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[COL_ITEM - 1]:
            continue
        item = str(row[COL_ITEM - 1]).strip()
        if not item or item == "None":
            continue

        requester = str(row[COL_REQUESTER - 1] or "").strip()
        date_val = row[COL_DATE_REQ - 1]
        status = str(row[COL_STATUS - 1] or "").strip()

        # 날짜 파싱
        date_str = ""
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        elif date_val:
            date_str = str(date_val)[:10]

        entry = {"item": item, "requester": requester, "date_requested": date_str, "status": status}

        # 미처리 주문
        if status in ("", "None") or "대기" in status or "미처리" in status:
            results["pending"].append(entry)

        # 최근 2주 주문
        if date_str:
            try:
                req_date = datetime.strptime(date_str[:10], "%Y-%m-%d")
                if req_date >= cutoff_2w:
                    results["recent_2weeks"].append(entry)
            except ValueError:
                pass

    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[Orders] done: pending={len(results['pending'])}, recent_2w={len(results['recent_2weeks'])}")
    print(f"saved: {OUTPUT_FILE}")

if __name__ == "__main__":
    collect()
